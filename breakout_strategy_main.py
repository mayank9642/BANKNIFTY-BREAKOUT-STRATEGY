# BANKNIFTY 5-min Breakout Strategy (Clean Implementation)
import time
import logging
import os
import shutil
import pytz
import threading
from datetime import datetime, timedelta
from src.config import load_config
from src.token_helper import ensure_valid_token
from src.fyers_api_utils import get_fyers_client, start_market_data_websocket, get_ltp, get_ltp_batch
from src.data_fetcher import DataFetcher
from src.symbol_formatter import convert_option_symbol_format, generate_option_symbol

class ISTFormatter(logging.Formatter):
    def __init__(self, fmt=None, datefmt=None):
        super().__init__(fmt=fmt, datefmt=datefmt)
        self.ist = pytz.timezone('Asia/Kolkata')
    def formatTime(self, record, datefmt=None):
        dt = datetime.fromtimestamp(record.created, self.ist)
        if datefmt:
            s = dt.strftime(datefmt)
        else:
            s = dt.strftime("%Y-%m-%d %H:%M:%S")
        return s

os.makedirs('logs', exist_ok=True)
log_file = 'logs/strategy.log'
def archive_strategy_log():
    if os.path.exists(log_file) and os.path.getsize(log_file) > 0:
        dt_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        archive_name = f"logs/strategy_{dt_str}.log"
        shutil.copy2(log_file, archive_name)
        # Clear log file safely to avoid null entries
        open(log_file, 'w').close()
        # Remove and recreate file handler to avoid writing to a corrupted file
        root_logger = logging.getLogger()
        for handler in root_logger.handlers[:]:
            if isinstance(handler, logging.FileHandler) and handler.baseFilename == os.path.abspath(log_file):
                root_logger.removeHandler(handler)
        file_handler = logging.FileHandler(log_file, mode='a')
        file_handler.setFormatter(ISTFormatter(fmt='%(asctime)s - %(levelname)s - %(message)s'))
        root_logger.addHandler(file_handler)
log_fmt = '%(asctime)s - %(levelname)s - %(message)s'
ist_formatter = ISTFormatter(fmt=log_fmt)

# Disable all existing loggers first
logging.disable(logging.DEBUG)

root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)

for handler in root_logger.handlers[:]:
    root_logger.removeHandler(handler)

file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')  # Overwrite log file each run
file_handler.setFormatter(ist_formatter)
file_handler.setLevel(logging.INFO)
root_logger.addHandler(file_handler)

# Use a UTF-8 console stream to avoid UnicodeEncodeError on Windows terminals
try:
    import sys, io
    utf8_stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    console_handler = logging.StreamHandler(utf8_stdout)
except Exception:
    console_handler = logging.StreamHandler()
console_handler.setFormatter(ist_formatter)
console_handler.setLevel(logging.INFO)
root_logger.addHandler(console_handler)

class Breakout5MinStrategy:
    def print_paper_trading_summary(self):
        """Prints a detailed summary of all paper trades/orders for the session, OCO logic, and P&L."""
        self.log_info("")
        self.log_info("=" * 79)
        self.log_info("PAPER TRADING SUMMARY - Bracket Order OCO Test")
        self.log_info("=" * 79)
        for order_id, order in self.paper_orders.items():
            self.log_info("")
            self.log_info(f"Order ID: {order_id}")
            self.log_info(f"  Symbol: {order.get('symbol','')}")
            self.log_info(f"  Status: {order.get('status','')}")
            self.log_info(f"  Entry Limit: {order.get('entry_limit','')}")
            self.log_info(f"  SL: {order.get('sl','')} | Target: {order.get('target','')}")
            self.log_info(f"  Qty: {order.get('qty',35)}")
            if order.get('status') == 'CANCELLED':
                self.log_info("  Cancelled (OCO - other leg filled)")
            if order.get('status') == 'FILLED':
                filled_time = order.get('filled_time', '')
                fill_price = order.get('entry_limit', '')
                self.log_info(f"  Filled at: {filled_time} @ {fill_price}")
                max_up = order.get('max_up_pnl', 0)
                max_dn = order.get('max_down_pnl', 0)
                entry = order.get('entry_limit', 0)
                qty = order.get('qty', 35)
                max_up_pct = (max_up / (entry * qty) * 100) if entry and qty else 0
                max_dn_pct = (max_dn / (entry * qty) * 100) if entry and qty else 0
                self.log_info(f"  Max Up: {max_up:.2f} ({max_up_pct:.2f}%) | Max Down: {max_dn:.2f} ({max_dn_pct:.2f}%)")
        self.log_info("")
        self.log_info("=" * 79)
        # OCO Test Result
        filled = [o for o in self.paper_orders.values() if o.get('status') == 'FILLED']
        cancelled = [o for o in self.paper_orders.values() if o.get('status') == 'CANCELLED']
        if filled and cancelled:
            self.log_info("OCO Test Result:")
            self.log_info("[OK] OCO LOGIC WORKING CORRECTLY!")
            self.log_info(f"   One order filled: {filled[0].get('symbol')}")
            self.log_info(f"   Other order cancelled: {cancelled[0].get('symbol')}")
        self.log_info("=" * 79)
        self.log_info("")
    def monitor_filled_paper_order(self, order_id):
        """Monitor paper trade using live prices (get_ltp) after breakout: logs entry, LTP, SL, target, P&L, max up/down, trailing SL, and final exit/performance logs."""
        import time
        from datetime import datetime, time as dtime
        order = self.paper_orders.get(order_id, {})
        symbol = order.get('symbol', 'UNKNOWN')
        entry_price = order.get('entry_limit', 0)
        qty = order.get('qty', 35)
        # --- Dynamic SL/Target based on VIX and premium ---
        vix = self.get_current_vix() if hasattr(self, 'get_current_vix') else 11  # fallback to 11 if not available
        self.log_info(f"[DEBUG] VIX value used at entry: {vix}")
        if vix is None:
            vix = 11
        if vix < 10:
            target_pct = 0.07 if entry_price <= 500 else 0.04
        elif vix < 12:
            target_pct = 0.10 if entry_price <= 500 else 0.05
        else:
            target_pct = 0.12 if entry_price <= 500 else 0.07
        sl_pct = target_pct
        sl = entry_price * (1 - sl_pct)
        target = entry_price * (1 + target_pct)
        trailing_sl = sl
        max_up_pnl = float('-inf')
        max_down_pnl = float('inf')
        exit_price = None
        exit_reason = None
        trade_active = True
        filled_dt = datetime.now(self.ist)
        filled_time = filled_dt.strftime('%H:%M:%S')
        order['filled_time'] = filled_time
        poll_interval_sec = 2
        market_close_time = dtime(15, 30)
        max_monitor_minutes = self.config.get('strategy', {}).get('paper_monitor_max_minutes', 120)
        max_invalid_ltp_streak = self.config.get('strategy', {}).get('max_invalid_ltp_streak', 20)
        invalid_ltp_streak = 0

        while trade_active:
            now_dt = datetime.now(self.ist)
            if now_dt.time() >= market_close_time:
                self.log_info(f"[PAPER EXIT] Market close reached for {symbol}. Exiting monitor loop.")
                exit_price = entry_price if exit_price is None else exit_price
                exit_reason = 'MARKET_CLOSE'
                break

            elapsed_min = (now_dt - filled_dt).total_seconds() / 60.0
            if elapsed_min >= max_monitor_minutes:
                self.log_info(f"[PAPER EXIT] Max monitor time reached ({max_monitor_minutes} min) for {symbol}. Exiting monitor loop.")
                exit_price = entry_price if exit_price is None else exit_price
                exit_reason = 'TIME_EXIT'
                break

            raw_ltp = self.get_ltp(symbol)
            try:
                ltp = float(raw_ltp)
                if ltp <= 0:
                    raise ValueError("Non-positive LTP")
                invalid_ltp_streak = 0
            except Exception:
                invalid_ltp_streak += 1
                self.log_info(
                    f"[WARN] Invalid LTP for {symbol} (streak {invalid_ltp_streak}/{max_invalid_ltp_streak}). "
                    f"Raw={raw_ltp}. Retrying..."
                )
                if invalid_ltp_streak >= max_invalid_ltp_streak:
                    self.log_info(f"[PAPER EXIT] Too many invalid LTP reads for {symbol}. Exiting monitor loop.")
                    exit_price = entry_price
                    exit_reason = 'DATA_EXIT'
                    break
                time.sleep(poll_interval_sec)
                continue
            pnl = (ltp - entry_price) * qty
            max_up_pnl = max(max_up_pnl, pnl)
            max_down_pnl = min(max_down_pnl, pnl)
            pnl_pct = (pnl / (entry_price * qty)) * 100 if entry_price else 0
            max_up_pct = (max_up_pnl / (entry_price * qty)) * 100 if entry_price else 0
            max_down_pct = (max_down_pnl / (entry_price * qty)) * 100 if entry_price else 0
            self.log_info(f"[PAPER STATUS] {symbol} | LTP: {ltp:.2f} | Entry: {entry_price:.2f} | SL: {sl:.2f} | Target: {target:.2f} | PnL: {pnl:.2f} | MaxUp: {max_up_pnl:.2f} ({max_up_pct:.2f}%) | MaxDn: {max_down_pnl:.2f} ({max_down_pct:.2f}%)")
            # Exit on SL/Target hit
            if ltp <= sl:
                self.log_info(f"[PAPER EXIT] Stop Loss hit at {ltp:.2f}")
                exit_price = ltp
                trade_active = False
                exit_reason = 'SL'
            elif ltp >= target:
                self.log_info(f"[PAPER EXIT] Target hit at {ltp:.2f}")
                exit_price = ltp
                trade_active = False
                exit_reason = 'TARGET'
            time.sleep(poll_interval_sec)  # Poll every 2 seconds (30 calls/min) — Fyers limit: 200/min

        if exit_price is None:
            exit_price = entry_price
        if not exit_reason:
            exit_reason = 'TIME_EXIT'
        # Update order with exit details
        order['exit_price'] = exit_price
        order['max_up_pnl'] = max_up_pnl
        order['max_down_pnl'] = max_down_pnl
        order['sl'] = sl
        order['target'] = target
        order['trailing_sl'] = trailing_sl
        order['qty'] = qty
        self.paper_orders[order_id] = order
        # Log trade completion and update balance/stats using the unified function (includes brokerage deduction)
        trade_pnl = (exit_price - entry_price) * qty if exit_price is not None else 0
        self.update_balance_on_trade_completion(pnl=trade_pnl, qty=qty, symbol=symbol)

        # For Excel logging, net_pnl is the raw trade P&L (before brokerage, as expected by log_paper_trade_to_excel)
        net_pnl = trade_pnl
        self.log_paper_trade_to_excel(
            order,
            symbol,
            entry_price,
            exit_price,
            qty,
            net_pnl,
            exit_reason
        )
    # (Removed duplicate call to log_paper_trade_to_excel outside function body)
    def __init__(self, simulation=False, paper_trading=False):
        self.simulation = simulation
        self.paper_trading = paper_trading
        self.config = load_config()
        self.fyers = get_fyers_client() if not simulation or paper_trading else None
        self.logger = logging.getLogger()
        self.ist = pytz.timezone('Asia/Kolkata')
        self.banknifty_symbol = self.config.get('strategy', {}).get('banknifty_symbol', 'NSE:NIFTYBANK-INDEX')
        self.banknifty_qty = self.config.get('strategy', {}).get('banknifty_qty', 35)
        self.breakout_buffer = self.config.get('strategy', {}).get('breakout_buffer', 5)
        self.log_file = 'logs/trade_history.csv'
        self.live_prices = {}
        self.data_socket = None
        # Initialize DataFetcher if we have a Fyers client (used for live or paper runs)
        self.data_fetcher = DataFetcher(self.fyers) if self.fyers is not None else None
        self.trade_executed_today = False
        self.trade_date = None
        # VIX config
        strat_cfg = self.config.get('strategy', {})
        self.vix_logging_enabled = strat_cfg.get('vix_logging_enabled', True)
        self.vix_symbol = strat_cfg.get('vix_symbol', 'NSE:INDIAVIX-INDEX')
        self.vix_log_csv = strat_cfg.get('vix_log_csv', 'logs/vix_levels.csv')
        
        # Paper trading: Simulate order state for BO OCO testing
        self.paper_orders = {}  # {order_id: {'symbol': str, 'status': str, 'entry_limit': float, 'placed_at': timestamp}}
        self._paper_order_seq = 0  # ensure unique IDs for paper orders
        
        # Capital Management: Track running balance
        self.initial_balance = self.config.get('simulation', {}).get('initial_balance', 100000)
        self.current_balance = self.initial_balance
        self.total_trades = 0
        self.winning_trades = 0
        self.losing_trades = 0
        self.total_profit_loss = 0.0
        
        # Load existing balance from file if it exists
        self.balance_file = 'logs/capital_balance.txt'

        self.load_current_balance()

        # Register atexit and signal handlers to always save balance on exit
        import atexit, signal
        def _save_balance_on_exit(*args, **kwargs):
            try:
                self.save_current_balance()
            except Exception:
                pass
        atexit.register(_save_balance_on_exit)
        for sig in (signal.SIGINT, signal.SIGTERM):
            try:
                signal.signal(sig, lambda signum, frame: (_save_balance_on_exit(), exit(0)))
            except Exception:
                pass

    def log_info(self, msg):
        # Replace emojis with ASCII text and remove non-ASCII characters to avoid UnicodeEncodeError
        import re
        emoji_map = {
            '🎯': '[TARGET]',
            '💰': '[BALANCE]',
            '📈': '[PERF]',
            '📊': '[EXCEL]',
            '✅': '[WIN]',
            '❌': '[LOSS]',
            '📉': '[DOWN]',
            '⚠️': '[WARN]',
        }
        for emoji, replacement in emoji_map.items():
            msg = msg.replace(emoji, replacement)
        # Remove any other non-ASCII characters
        safe_msg = re.sub(r'[^\x00-\x7F]+', '', msg)
        self.logger.info(safe_msg)
    
    def load_current_balance(self):
        """Load current balance from file if it exists"""
        try:
            if os.path.exists(self.balance_file):
                with open(self.balance_file, 'r') as f:
                    lines = f.readlines()
                    if len(lines) >= 6:
                        self.current_balance = float(lines[0].strip())
                        self.total_trades = int(lines[1].strip())
                        self.winning_trades = int(lines[2].strip())
                        self.losing_trades = int(lines[3].strip())
                        self.total_profit_loss = float(lines[4].strip())
                        
                        self.log_info(f"💰 Loaded existing balance: ₹{self.current_balance:,.2f}")
                        self.log_info(f"📊 Stats: {self.total_trades} trades ({self.winning_trades}W/{self.losing_trades}L) | Net P&L: ₹{self.total_profit_loss:,.2f}")
                    else:
                        self.log_info(f"💰 Starting fresh with initial balance: ₹{self.initial_balance:,.2f}")
            else:
                self.log_info(f"💰 Starting fresh with initial balance: ₹{self.initial_balance:,.2f}")
        except Exception as e:
            self.log_info(f"⚠️ Error loading balance: {e}. Using initial balance: ₹{self.initial_balance:,.2f}")
    
    def save_current_balance(self):
        """Save current balance to file"""
        try:
            os.makedirs('logs', exist_ok=True)
            with open(self.balance_file, 'w') as f:
                f.write(f"{self.current_balance}\n")
                f.write(f"{self.total_trades}\n")
                f.write(f"{self.winning_trades}\n")
                f.write(f"{self.losing_trades}\n")
                f.write(f"{self.total_profit_loss}\n")
                f.write(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        except Exception as e:
            self.log_info(f"⚠️ Error saving balance: {e}")
    
    def update_balance_on_trade_completion(self, pnl, qty, symbol):
        """Update balance when a trade is completed"""
        try:
            # Calculate net P&L (P&L minus brokerage)
            brokerage_cost = 50  # Fixed brokerage per trade (buy+sell)
            net_pnl = pnl - brokerage_cost
            
            self.total_trades += 1
            self.total_profit_loss += net_pnl
            self.current_balance += net_pnl
            
            # Save the updated balance
            self.save_current_balance()
            
            if net_pnl > 0:
                self.winning_trades += 1
                status = "PROFIT ✅"
            else:
                self.losing_trades += 1
                status = "LOSS ❌"
            
            win_rate = (self.winning_trades / self.total_trades * 100) if self.total_trades > 0 else 0
            
            self.log_info(f"🎯 Trade #{self.total_trades} completed: {symbol} | Net P&L: ₹{net_pnl:,.2f} ({status})")
            self.log_info(f"💰 Updated Balance: ₹{self.current_balance:,.2f} (Change: ₹{pnl:+,.2f})")
            self.log_info(f"📈 Performance: {self.winning_trades}W/{self.losing_trades}L | Win Rate: {win_rate:.1f}% | Net P&L: ₹{self.total_profit_loss:+,.2f}")
            
            # Save to file
            self.save_current_balance()
            
        except Exception as e:
            self.log_info(f"⚠️ Error updating balance: {e}")
    
    def reset_balance_to_initial(self):
        """Reset balance to initial amount (₹100,000)"""
        try:
            self.current_balance = self.initial_balance
            self.total_trades = 0
            self.winning_trades = 0
            self.losing_trades = 0
            self.total_profit_loss = 0.0
            
            self.save_current_balance()
            self.log_info(f"🔄 Balance reset to initial amount: ₹{self.initial_balance:,.2f}")
            
        except Exception as e:
            self.log_info(f"⚠️ Error resetting balance: {e}")

    def log_paper_trade_to_excel(self, order, symbol, entry_price, exit_price, qty, pnl, exit_reason):
        """Log paper trading results to Excel and CSV files"""
        try:
            from datetime import datetime
            
            # Prepare trade data for Excel
            entry_time = datetime.fromtimestamp(order.get('placed_at', 0), self.ist).strftime('%Y-%m-%d %H:%M:%S') if 'placed_at' in order else ''
            exit_time = datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S')
            
            # Get order details
            sl = order.get('sl', 0)
            target = order.get('target', 0)
            
            # Get max up/down PnL values - handle cases where they might not be set properly
            max_up_pnl = order.get('max_up_pnl', 0)
            max_down_pnl = order.get('max_down_pnl', 0)
            
            # If max values are still at their initial values (inf/-inf), use PnL instead
            if max_up_pnl == float('-inf') or max_up_pnl == 0:
                max_up_pnl = max(0, pnl) if pnl is not None else 0
            if max_down_pnl == float('inf') or max_down_pnl == 0:
                max_down_pnl = min(0, pnl) if pnl is not None else 0
            
            # Calculate percentages based on the corrected max values
            total_investment = entry_price * qty
            max_up_pct = (max_up_pnl / total_investment * 100) if total_investment and max_up_pnl else 0
            max_down_pct = (max_down_pnl / total_investment * 100) if total_investment and max_down_pnl else 0
            
            # Use single Forward Testing file that appends all trades
            excel_file = 'logs/Forward Testing Trade History.xlsx'
            csv_file = 'logs/Forward Testing Trade History.csv'
            
            status_columns = [
                'Entry DateTime', 'Index', 'Symbol', 'Direction', 'Entry Price', 
                'Exit DateTime', 'Exit Price', 'Stop Loss', 'Target', 'Trailing SL', 
                'Quantity', 'Brokerage', 'P&L', 'Net P&L', 'Margin Required', '% Gain/Loss', 
                'Max Up (₹)', 'Max Down (₹)', 'Max Up (%)', 'Max Down (%)', 'VIX', 'Balance After Trade'
            ]
            
            # Calculate proper values for Excel
            total_investment = entry_price * qty
            pnl_percentage = (pnl / total_investment * 100) if total_investment else 0
            lots_traded = qty / 35  # Convert quantity to lots (35 qty = 1 lot)
            brokerage_cost = lots_traded * 50  # ₹50 per lot for buy and sell combined
            margin_required = total_investment  # Total investment IS the margin required
            trailing_sl_value = sl  # Use actual trailing SL if available
            
            # Calculate Net P&L (P&L minus brokerage)
            net_pnl = pnl - brokerage_cost if pnl is not None else -brokerage_cost
            
            # Balance is already updated by update_balance_on_trade_completion()
            # Just get the current balance for Excel logging
            balance_after_trade = self.current_balance
            
            # Get current VIX level
            vix_value = self.get_current_vix() if hasattr(self, 'get_current_vix') else order.get('vix', 0)
            
            final_row = [
                entry_time, 'BANKNIFTY', symbol, 'BUY', entry_price,
                exit_time, exit_price, sl, target, trailing_sl_value, 
                qty, brokerage_cost, pnl, net_pnl, margin_required, pnl_percentage,
                max_up_pnl, max_down_pnl, max_up_pct, max_down_pct, vix_value, balance_after_trade
            ]
            
            # Write to Excel and CSV
            self._append_final_row_with_format(excel_file, csv_file, final_row, status_columns)
            
            # Also log to regular trade history
            self.log_trade(symbol, exit_price, qty, 'BUY', exit_reason, exit_time)
            
            self.log_info(f"📊 Trade logged to Excel: {symbol} | Entry: ₹{entry_price:.2f} | Exit: ₹{exit_price:.2f} | P&L: ₹{pnl:.2f}")
            
        except Exception as e:
            self.log_info(f"⚠️ Error logging paper trade to Excel: {e}")

    def run(self):
        archive_strategy_log()
        self.log_info('Starting 5-min breakout strategy (BANKNIFTY).')

        # --- Robust market open/holiday check ---
        now = datetime.now(self.ist)
        current_time = now.time()
        current_date = now.date()
        # Check if it's a weekend
        if now.weekday() >= 5:  # Saturday=5, Sunday=6
            self.log_info('Market is closed (weekend). Exiting strategy.')
            return
        # Check if it's a market holiday
        holidays = self.config.get('market_holidays', [])
        if current_date.strftime('%Y-%m-%d') in holidays:
            self.log_info('Market is closed (holiday). Exiting strategy.')
            return
        # Always wait for market open if before 09:15, regardless of mode
        market_open = datetime.strptime(self.config['timing']['market_open_time'], '%H:%M').time()
        if current_time < market_open:
            self.log_info(f'Market not open yet (current time: {current_time}). Waiting for market to open...')
        self.wait_for_market_open()
        # After market open, check if within trading hours
        now = datetime.now(self.ist)
        current_time = now.time()
        market_close = datetime.strptime(self.config['timing']['trading_end_time'], '%H:%M').time()
        if not (market_open <= current_time <= market_close):
            self.log_info(f'Market is closed (current time: {current_time}). Exiting strategy.')
            return

        self.wait_until_920()

        # ── VIX FILTER ───────────────────────────────────────────────────────
        # Based on historical analysis of real forward-testing trades:
        #
        #   VIX < 12  →  2 trades |  0% win | Net P&L: -2,788  ❌ AVOID (new)
        #   VIX 12–15 → sweet spot | 75–88% win rate           ✅ TRADE
        #   VIX 15–17 → 12 trades | 58% win | Cumul P&L: +3,605 ✅ TRADE
        #   VIX 17–19 → 21 trades | 43% win | Cumul P&L: -5,829 ❌ AVOID
        #   VIX 21–24 →  6 trades | 67% win | Cumul P&L: +6,974 ✅ TRADE
        #   VIX > 24  →  9 trades | 67% win | Cumul P&L: +6,712 ✅ TRADE
        #
        # LOW VIX (<12): market too calm, options don't move enough to hit
        # target but SL still gets clipped. 0% win rate in forward testing.
        # HIGH VIX 17-19: chop zone, premiums inflated but no clean direction.
        # ─────────────────────────────────────────────────────────────────────
        _vix_now = self.get_current_vix()
        if _vix_now is not None:
            if _vix_now < 12.0:
                self.log_info(
                    f'[VIX TOO LOW] India VIX = {_vix_now:.2f} (below 12.0). '
                    f'Forward test: 0% win rate below VIX 12. '
                    f'Market too calm for breakout strategy. Skipping trade today.'
                )
                print('\033[93m' + '=' * 70 + '\033[0m')
                print('\033[93m' + '  [VIX TOO LOW]   India VIX = {:.2f}'.format(_vix_now) + '\033[0m')
                print('\033[93m' + '  VIX below 12.0 -- market too calm for breakout              ' + '\033[0m')
                print('\033[93m' + '  Forward test: 0% win rate | 2/2 losses below this level     ' + '\033[0m')
                print('\033[93m' + '  NO TRADE TODAY -- Strategy exiting cleanly.                 ' + '\033[0m')
                print('\033[93m' + '=' * 70 + '\033[0m')
                return
            elif 17.0 <= _vix_now < 19.0:
                self.log_info(
                    f'[VIX RED ZONE] India VIX = {_vix_now:.2f} (Range 17.0-19.0 detected). '
                    f'Historical record: 21 trades, 43% win rate, Cumul P&L -5829. '
                    f'Skipping trade for today. Strategy ended cleanly.'
                )
                print('\033[91m' + '=' * 70 + '\033[0m')
                print('\033[91m' + '  [VIX RED ZONE]  India VIX = {:.2f}'.format(_vix_now) + '\033[0m')
                print('\033[91m' + '  VIX range 17.0 - 19.0 detected                              ' + '\033[0m')
                print('\033[91m' + '  Historical: 21 trades | 43% win rate | Cumul P&L: -5,829    ' + '\033[0m')
                print('\033[91m' + '  NO TRADE TODAY -- Strategy exiting cleanly.                 ' + '\033[0m')
                print('\033[91m' + '=' * 70 + '\033[0m')
                return
        else:
            self.log_info('[VIX] Could not fetch VIX for filter check -- proceeding with trade.')

        t = threading.Thread(target=self.monitor_index, args=(self.banknifty_symbol, self.banknifty_qty, 'BANKNIFTY'), daemon=True)
        t.start()
        t.join()

    def wait_for_market_open(self):
        # Skip only in pure simulation (backtesting). Paper trading must wait for real market open.
        if self.simulation and not self.paper_trading:
            self.log_info('[SIMULATION] Skipping market open wait.')
            return
        now = datetime.now(self.ist)
        market_open = now.replace(hour=9, minute=15, second=0, microsecond=0)
        while now < market_open:
            self.log_info('Waiting for market to open (09:15 IST)...')
            time.sleep(30)
            now = datetime.now(self.ist)

    def wait_until_920(self):
        # Skip only in pure simulation (backtesting). Paper trading must wait for 09:20 candle.
        if self.simulation and not self.paper_trading:
            self.log_info('[SIMULATION] Skipping 9:20 wait. Starting immediately.')
            return
        now = datetime.now(self.ist)
        target = now.replace(hour=9, minute=20, second=0, microsecond=0)
        if now < target:
            seconds = (target - now).total_seconds()
            self.log_info(f'Waiting {int(seconds)} seconds until 9:20 IST for first 5-min candle to form...')
            time.sleep(seconds)
        self.log_info('Reached 9:20 IST. Proceeding to fetch first 5-min candle.')
        # Log India VIX snapshot at 09:20
        self.log_vix_snapshot(label='09:20')

    def log_vix_snapshot(self, label='09:20'):
        """Fetch and log India VIX LTP and persist to CSV for later analysis."""
        try:
            if not self.vix_logging_enabled:
                return
            if not self.fyers:
                self.log_info('[VIX] Fyers client not initialized; skipping VIX snapshot.')
                return
            vix_ltp = self.get_ltp(self.vix_symbol)
            if vix_ltp is None:
                self.log_info(f"[VIX] Could not fetch VIX LTP for {self.vix_symbol}.")
                return
            ts = datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S')
            self.log_info(f"[VIX] {label} snapshot: {self.vix_symbol} LTP = {vix_ltp}")
            # Append to CSV
            try:
                os.makedirs(os.path.dirname(self.vix_log_csv), exist_ok=True)
                new_file = not os.path.exists(self.vix_log_csv)
                import csv
                with open(self.vix_log_csv, 'a', newline='', encoding='utf-8') as cf:
                    writer = csv.writer(cf)
                    if new_file:
                        writer.writerow(['Date', 'Time', 'Label', 'Symbol', 'VIX'])
                    date_str = datetime.now(self.ist).strftime('%Y-%m-%d')
                    time_str = datetime.now(self.ist).strftime('%H:%M:%S')
                    writer.writerow([date_str, time_str, label, self.vix_symbol, f"{vix_ltp}"])
            except Exception as e_csv:
                self.log_info(f"[VIX] Failed to write VIX CSV: {e_csv}")
        except Exception as e:
            self.log_info(f"[VIX] Error logging VIX snapshot: {e}")

    def fetch_5min_candle(self, symbol):
        # Always use DataFetcher for live candle data
        if self.data_fetcher is None:
            self.log_info(f"[ERROR] DataFetcher not initialized for {symbol}.")
            return None
        candle_data = self.data_fetcher.get_first_5min_candle(symbol)
        if candle_data:
            o, h, l, cl, candle_time = candle_data
            self.log_info(f"5-min OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
            return candle_data
        self.log_info(f"[ERROR] Could not fetch 5-min candle for {symbol}.")
        return None

    def fetch_option_ohlc(self, symbol):
        # Always use DataFetcher for live option OHLC data
        if self.data_fetcher is None:
            self.log_info(f"[ERROR] DataFetcher not initialized for option {symbol}.")
            return None
        candle_data = self.data_fetcher.get_first_5min_candle(symbol)
        if candle_data:
            o, h, l, cl, candle_time = candle_data
            self.log_info(f"5-min OHLC for option {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
            return candle_data
        self.log_info(f"[ERROR] Could not fetch 5-min candle for option {symbol}.")
        return None

    def get_atm_option_symbol(self, spot, option_type, index_name):
        """Generate ATM option symbol with proper expiry selection (BANKNIFTY: Fyers option chain only)"""
        try:
            # Check if BANKNIFTY options are enabled
            if 'BANK' in index_name.upper():
                banknifty_options_enabled = self.config.get('strategy', {}).get('banknifty_options_enabled', False)
                if not banknifty_options_enabled:
                    self.log_info(f"BANKNIFTY options not enabled in config - skipping {option_type} for {index_name}")
                    return None
            # Calculate step size and ATM strike
            step = 100 if 'BANK' in index_name.upper() else 50
            strike = round(spot / step) * step
            today = datetime.now(self.ist)
            if 'BANK' in index_name.upper():
                # BANKNIFTY: Always use next available expiry from Fyers option chain
                from src.banknifty_symbol_helper import get_next_banknifty_expiry, get_banknifty_option_symbol
                expiry_date = get_next_banknifty_expiry(today)
                underlying = 'BANKNIFTY'
                try:
                    symbol = get_banknifty_option_symbol(int(strike), option_type, expiry_date.date())
                    self.log_info(f"Selected BANKNIFTY {option_type} symbol from Fyers option chain: {symbol}")
                    return symbol
                except Exception as e:
                    self.log_info(f"[ERROR] BANKNIFTY option chain lookup failed: {e}")
                    # Fallback to formatter if option chain fails
                    from src.symbol_formatter import generate_option_symbol
                    symbol = generate_option_symbol(underlying, expiry_date.date(), int(strike), option_type)
                    self.log_info(f"Fallback BANKNIFTY symbol: {symbol}")
                    return symbol
            else:
                # NIFTY: Weekly options expire on Thursday
                target_weekday = 3  # Thursday
                days_to_expiry = (target_weekday - today.weekday()) % 7
                if days_to_expiry == 0:  # Today is Thursday
                    if today.hour > 15 or (today.hour == 15 and today.minute >= 30):
                        days_to_expiry = 7  # Next Thursday after market close
                expiry_date = today + timedelta(days=days_to_expiry)
            # For NIFTY, use Fyers option chain to get exact symbol
            from src.nifty_symbol_helper import get_nifty_atm_option_symbol
            try:
                symbol = get_nifty_atm_option_symbol(spot, expiry_date.strftime('%d-%m-%Y'), option_type)
                if symbol:
                    self.log_info(f"Selected NIFTY {option_type} symbol from Fyers option chain: {symbol}")
                    return symbol
                else:
                    self.log_info(f"[ERROR] NIFTY option chain lookup failed, falling back to formatter.")
                    symbol = generate_option_symbol('NIFTY', expiry_date.date(), int(strike), option_type)
                    self.log_info(f"Fallback NIFTY symbol: {symbol}")
                    return symbol
            except Exception as e:
                self.log_info(f"[ERROR] NIFTY option chain lookup exception: {e}")
                symbol = generate_option_symbol('NIFTY', expiry_date.date(), int(strike), option_type)
                self.log_info(f"Fallback NIFTY symbol: {symbol}")
                return symbol
        except Exception as e:
            self.log_info(f"[ERROR] Failed to generate option symbol: {e}")
            return None

    def get_current_vix(self):
        """Get current VIX value"""
        try:
            if self.fyers:
                vix_symbol = "NSE:INDIAVIX-INDEX"
                response = self.fyers.quotes({"symbols": vix_symbol})
                if response and response.get('s') == 'ok':
                    return response['d'][0]['v']['lp'] if response.get('d') and len(response['d']) > 0 else 0
            self.log_info(f"[ERROR] Fyers client not initialized for VIX fetch.")
            return None
        except Exception as e:
            self.log_info(f"Error fetching VIX: {e}")
            return None

    def get_ltp(self, symbol):
        # Always use Fyers API utility for LTP
        from src.fyers_api_utils import get_ltp
        from datetime import datetime
        if self.fyers:
            ltp = get_ltp(self.fyers, symbol)
            return ltp
        else:
            self.log_info(f"[ERROR] Fyers client not initialized for LTP fetch.")
            return None

    def setup_websocket(self, symbols):
        # Dummy websocket setup
        pass

    def get_ltp_batch(self, symbols):
        """Fetch LTP for multiple symbols in ONE API call (Fyers supports up to 50 symbols).
        Returns dict: {symbol: float_ltp}. Missing symbols absent from dict.
        In pure simulation (backtesting) returns 100.0 per symbol without calling API.
        """
        if self.simulation and not self.paper_trading:
            return {s: 100.0 for s in symbols}
        # Prefer DataFetcher batch method (has retry + rate-limit handling)
        if self.data_fetcher:
            try:
                result = self.data_fetcher.get_ltp_batch(symbols)
                if result:
                    return result
            except Exception as e:
                self.log_info(f"[ERROR] DataFetcher batch LTP failed: {e}")
        # Fallback to fyers_api_utils batch function
        try:
            return get_ltp_batch(self.fyers, symbols)
        except Exception as e:
            self.log_info(f"Error in batch LTP for {symbols}: {e}")
            return {}

    def monitor_index(self, symbol, qty, index_name):
        candle = self.fetch_5min_candle(symbol)
        if not candle:
            return None
        open_, high, low, close, candle_time = candle
        ce_symbol = self.get_atm_option_symbol(high, 'CE', index_name)
        pe_symbol = self.get_atm_option_symbol(low, 'PE', index_name)
        ce_ohlc = self.fetch_option_ohlc(ce_symbol)
        pe_ohlc = self.fetch_option_ohlc(pe_symbol)
        if not ce_ohlc or not pe_ohlc:
            return None
        ce_high = ce_ohlc[1]
        pe_high = pe_ohlc[1]
        ce_close = ce_ohlc[3]
        pe_close = pe_ohlc[3]
        
        # Use configurable breakout buffer (default 2 points)
        breakout_buffer = self.config.get('strategy', {}).get('breakout_buffer_points', 2)  # Default 2 points
        ce_breakout = ce_high + breakout_buffer
        pe_breakout = pe_high + breakout_buffer
        self.log_info(f"Monitoring CE {ce_symbol} for breakout above {ce_breakout} (buffer: {breakout_buffer})")
        self.log_info(f"Monitoring PE {pe_symbol} for breakout above {pe_breakout} (buffer: {breakout_buffer})")
        self.monitor_option_high_breakout(ce_symbol, pe_symbol, ce_breakout, pe_breakout, qty, index_name, ce_close, pe_close)

    def monitor_option_high_breakout(self, ce_symbol, pe_symbol, ce_breakout, pe_breakout, qty, index_name, ce_close, pe_close):
        from datetime import datetime, timedelta, time as dtime
        breakout_taken = False
        # Monitor until 3:30 PM IST
        market_close_time = dtime(15, 30)
        # OCO Implementation: Place both CE and PE bracket orders IMMEDIATELY
        ce_order_id = None
        pe_order_id = None
        oco_entry_taken = False  # Track if one of the OCO orders filled
        # Calculate entry prices at breakout levels
        ce_entry_price = ce_breakout
        pe_entry_price = pe_breakout
        quantity = qty * 35  # Convert lots to quantity (35 units per lot for BANKNIFTY)
        self.log_info("=" * 70)
        self.log_info("PLACING OCO BRACKET ORDERS AT BREAKOUT LEVELS")
        self.log_info("=" * 70)
        # Place BOTH bracket orders immediately
        self.log_info(f"Placing CE BO: {ce_symbol} @ {ce_entry_price:.2f} (trigger when price >= {ce_breakout:.2f})")
        ce_order_id = self.place_bracket_order(ce_symbol, ce_entry_price, quantity, ce_breakout, index_name)
        # Ensure paper_orders is populated for simulation/paper mode
        if self.simulation or self.paper_trading:
            self.paper_orders[ce_order_id] = {
                'symbol': ce_symbol,
                'status': 'PENDING',
                'entry_limit': ce_entry_price,
                'placed_at': time.time()
            }
        self.log_info(f"Placing PE BO: {pe_symbol} @ {pe_entry_price:.2f} (trigger when price >= {pe_breakout:.2f})")
        pe_order_id = self.place_bracket_order(pe_symbol, pe_entry_price, quantity, pe_breakout, index_name)
        if self.simulation or self.paper_trading:
            self.paper_orders[pe_order_id] = {
                'symbol': pe_symbol,
                'status': 'PENDING',
                'entry_limit': pe_entry_price,
                'placed_at': time.time()
            }
        if not ce_order_id or not pe_order_id:
            self.log_info("[ERROR] Failed to place one or both bracket orders. Stopping strategy.")
            return
        self.log_info("=" * 70)
        self.log_info("Both OCO bracket orders placed successfully!")
        self.log_info(f"CE Order ID: {ce_order_id}")
        self.log_info(f"PE Order ID: {pe_order_id}")
        self.log_info("Orders are now at broker with SL/TP configured.")
        self.log_info("Monitoring order status... whichever triggers first will cancel the other.")
        self.log_info("=" * 70)
        # ==== PHASE 2: Monitor order status and implement OCO cancellation ====
        while not oco_entry_taken:
                now = datetime.now(self.ist).time()
                if now >= market_close_time:
                    self.log_info("[INFO] Market close reached (3:30 PM). Stopping breakout monitoring.")
                    break
                # --- FILL SIMULATION LOGIC ---
                # Batch fetch CE+PE in ONE API call (halves quota usage vs two separate calls)
                _batch = self.get_ltp_batch([ce_symbol, pe_symbol])
                ce_ltp = _batch.get(ce_symbol)
                pe_ltp = _batch.get(pe_symbol)
                self.log_info(f"[DEBUG] Breakout levels: CE={ce_breakout}, PE={pe_breakout} | LTPs: CE={ce_ltp}, PE={pe_ltp}")
                ce_ltp_rounded = round(float(ce_ltp), 2) if ce_ltp is not None else None
                pe_ltp_rounded = round(float(pe_ltp), 2) if pe_ltp is not None else None
                ce_breakout_rounded = round(float(ce_breakout), 2)
                pe_breakout_rounded = round(float(pe_breakout), 2)
                ce_status = self.get_order_status(ce_order_id)
                pe_status = self.get_order_status(pe_order_id)
                if ce_status == "PENDING" and ce_ltp_rounded is not None and ce_ltp_rounded >= ce_breakout_rounded:
                    self.log_info(f"[DEBUG] CE LTP {ce_ltp_rounded} >= CE breakout {ce_breakout_rounded} -- triggering fill logic.")
                    self.log_info(f"[DEBUG] paper_orders keys: {list(self.paper_orders.keys())}")
                    if ce_order_id in self.paper_orders and pe_order_id in self.paper_orders:
                        self.paper_orders[ce_order_id]["status"] = "FILLED"
                        self.paper_orders[pe_order_id]["status"] = "CANCELLED"
                        ce_status = "FILLED"
                        pe_status = "CANCELLED"
                        self.log_info(f"[SIMULATION] CE breakout triggered! CE order FILLED at {ce_ltp_rounded}. PE order CANCELLED (OCO logic).")
                    else:
                        self.log_info(f"[ERROR] Order ID(s) not found in paper_orders: CE={ce_order_id}, PE={pe_order_id}")
                elif pe_status == "PENDING" and pe_ltp_rounded is not None and pe_ltp_rounded >= pe_breakout_rounded:
                    self.log_info(f"[DEBUG] PE LTP {pe_ltp_rounded} >= PE breakout {pe_breakout_rounded} -- triggering fill logic.")
                    self.log_info(f"[DEBUG] paper_orders keys: {list(self.paper_orders.keys())}")
                    if pe_order_id in self.paper_orders and ce_order_id in self.paper_orders:
                        self.paper_orders[pe_order_id]["status"] = "FILLED"
                        self.paper_orders[ce_order_id]["status"] = "CANCELLED"
                        pe_status = "FILLED"
                        ce_status = "CANCELLED"
                        self.log_info(f"[SIMULATION] PE breakout triggered! PE order FILLED at {pe_ltp_rounded}. CE order CANCELLED (OCO logic).")
                    else:
                        self.log_info(f"[ERROR] Order ID(s) not found in paper_orders: CE={ce_order_id}, PE={pe_order_id}")
                self.log_info(f"Order Status: CE={ce_status} | PE={pe_status}")
                # Check if CE triggered/filled
                if ce_status in ["FILLED", "TRIGGERED"]:
                    self.log_info("=" * 70)
                    self.log_info(f"CE ORDER TRIGGERED/FILLED! Cancelling PE order...")
                    self.log_info("=" * 70)
                    # Cancel PE order if still pending
                    if pe_status in ["PENDING", "OPEN"]:
                        cancel_success = self.cancel_order(pe_order_id, pe_symbol)
                        if cancel_success:
                            self.log_info(f"PE order cancelled successfully")
                        else:
                            self.log_info(f"[WARNING] Failed to cancel PE order - it may have already triggered")
                    self.log_info(f"CE position active with automatic SL and Target management by broker")
                    oco_entry_taken = True
                    self.trade_executed_today = True
                    # After entry, monitor the filled paper position until SL/Target/timeout
                    if (self.paper_trading or self.simulation):
                        self.monitor_filled_paper_order(ce_order_id)
                    break
                # Check if PE triggered/filled
                elif pe_status in ["FILLED", "TRIGGERED"]:
                    self.log_info("=" * 70)
                    self.log_info(f"PE ORDER TRIGGERED/FILLED! Cancelling CE order...")
                    self.log_info("=" * 70)
                    # Cancel CE order if still pending
                    if ce_status in ["PENDING", "OPEN"]:
                        cancel_success = self.cancel_order(ce_order_id, ce_symbol)
                        if cancel_success:
                            self.log_info(f"CE order cancelled successfully")
                        else:
                            self.log_info(f"[WARNING] Failed to cancel CE order - it may have already triggered")
                    self.log_info(f"PE position active with automatic SL and Target management by broker")
                    oco_entry_taken = True
                    self.trade_executed_today = True
                    if (self.paper_trading or self.simulation):
                        self.monitor_filled_paper_order(pe_order_id)
                    break
                # Check if both orders failed/cancelled
                elif ce_status in ["CANCELLED", "REJECTED"] and pe_status in ["CANCELLED", "REJECTED"]:
                    self.log_info("[ERROR] Both orders failed or were cancelled. Stopping strategy.")
                    break
                # Optional: Fetch and log current LTP for monitoring (informational only)
                if ce_ltp and pe_ltp:
                    self.log_info(f"Current Prices: CE LTP: {ce_ltp:.2f} | PE LTP: {pe_ltp:.2f}")
                time.sleep(2)  # Poll every 2s (60 calls/min for CE+PE) — Fyers limit: 200/min
        if not oco_entry_taken:
            self.log_info(f"No entry taken within monitoring window.")
        
        # Paper trading summary
        if (self.paper_trading or self.simulation) and self.paper_orders:
            self.print_paper_trading_summary()

    def wait_for_market_open(self):
        now = datetime.now(self.ist)
        market_open = now.replace(hour=9, minute=15, second=0, microsecond=0)
        if now >= market_open:
            # After 09:15, but always want to use 09:15-09:20 window for breakout
            self.skip_time_rule = False  # Always use 09:15-09:20
            self.log_info("[USER] Script started after 09:15. Will attempt to fetch 09:15-09:20 OHLC for breakout.")
            return
        while True:
            now = datetime.now(self.ist).time()
            if now >= datetime.strptime('09:15', '%H:%M').time():
                break
            self.log_info('Waiting for market to open (09:15 IST)...')
            time.sleep(30)

    def wait_until_920(self):
        now = datetime.now(self.ist)
        target = now.replace(hour=9, minute=20, second=0, microsecond=0)
        if now < target:
            seconds = (target - now).total_seconds()
            self.log_info(f'Waiting {int(seconds)} seconds until 9:20 IST for first 5-min candle to form...')
            time.sleep(seconds)
            self.log_info('Reached 9:20 IST. Proceeding to fetch first 5-min candle.')

    def collect_live_5min_ohlc(self, symbol):
        # Deprecated: No longer used. Always use Fyers historical API for index OHLC.
        self.log_info(f"[SKIP] collect_live_5min_ohlc is disabled. Using Fyers historical API only for {symbol}.")
        return None

    def fetch_5min_candle(self, symbol):
        # Always use Fyers historical API to fetch 5-min OHLC for index at 09:20 IST
        if self.simulation and not self.paper_trading:
            return (20000, 20020, 19980, 20010, '09:15')
            
        # Use the new DataFetcher for more reliable candle data
        candle_data = self.data_fetcher.get_first_5min_candle(symbol)
        if candle_data:
            o, h, l, cl, candle_time = candle_data
            self.log_info(f"[ENHANCED] 5-min OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
            return candle_data
            
        # If DataFetcher fails, fall back to the old method
        self.log_info(f"[FALLBACK] DataFetcher failed for {symbol}, trying old method")
        
        from datetime import datetime, timedelta
        ist = pytz.timezone('Asia/Kolkata')
        now = datetime.now(ist)
        # Target 09:15-09:20 candle for breakout
        target_time = now.replace(hour=9, minute=20, second=0, microsecond=0)
        if now < target_time:
            # If before 9:20, wait until 9:20
            wait_sec = (target_time - now).total_seconds()
            if wait_sec > 0:
                self.log_info(f"Waiting {int(wait_sec)} seconds until 9:20 IST for first 5-min candle to form...")
                time.sleep(wait_sec)
            now = datetime.now(ist)
        # Retry logic for Fyers API
        for attempt in range(5):
            try:
                from_time = target_time - timedelta(minutes=5)
                to_time = target_time
                range_from = int(from_time.timestamp())
                range_to = int(to_time.timestamp())
                data = {
                    "symbol": symbol,
                    "resolution": "5",
                    "date_format": "0",  # Use 0 for epoch timestamps
                    "range_from": range_from,
                    "range_to": range_to,
                    "cont_flag": "1"
                }
                candles = self.fyers.history(data)
                if candles.get('s') == 'ok' and candles.get('candles'):
                    c = candles['candles'][-1]
                    o, h, l, cl = c[1], c[2], c[3], c[4]
                    candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                    self.log_info(f"[API] 5-min OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                    return (o, h, l, cl, candle_time)
                else:
                    self.log_info(f"[RETRY] No 5-min candle data returned for {symbol} from Fyers (attempt {attempt+1}/5). Retrying...")
                    time.sleep(2)
            except Exception as e:
                self.log_info(f"[RETRY] Error fetching 5-min candle for {symbol} (attempt {attempt+1}/5): {e}")
                time.sleep(2)
        # Fallback: try to fetch the most recent 5-min candle for today
        try:
            today = now.replace(hour=0, minute=0, second=0, microsecond=0)
            data = {
                "symbol": symbol,
                "resolution": "5",
                "date_format": "0",
                "range_from": int(today.timestamp()),
                "range_to": int(now.timestamp()),
                "cont_flag": "1"
            }
            candles = self.fyers.history(data)
            if candles.get('s') == 'ok' and candles.get('candles'):
                c = candles['candles'][-1]
                o, h, l, cl = c[1], c[2], c[3], c[4]
                candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                self.log_info(f"[FALLBACK-LAST] Using most recent 5-min candle for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                return (o, h, l, cl, candle_time)
            else:
                self.log_info(f"[ERROR] No 5-min candle data available for {symbol} even after all fallbacks.")
                return None
        except Exception as e:
            self.log_info(f"[ERROR] Final fallback error fetching 5-min candle for {symbol}: {e}")
            return None

    def fetch_option_ohlc(self, symbol):
        # Fetch 5-min OHLC for option symbol using Fyers historical API with retry/fallback
        if self.simulation and not self.paper_trading:
            return (100, 106, 99, 105, '09:20')
            
        # Use the new DataFetcher for more reliable option data
        if self.data_fetcher:
            try:
                candle_data = self.data_fetcher.get_first_5min_candle(symbol)
                if candle_data:
                    o, h, l, cl, candle_time = candle_data
                    self.log_info(f"[ENHANCED] 5-min option OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                    return candle_data
            except Exception as e:
                self.log_info(f"[ERROR] DataFetcher failed for option {symbol}: {e}")
        
        # If DataFetcher fails, fall back to the old method
        ist = pytz.timezone('Asia/Kolkata')
        now = datetime.now(ist)
        # Try up to 5 times with 2s delay to allow for Fyers data lag
        for attempt in range(5):
            try:
                # Try to get the 09:15-09:20 candle, else fallback to latest available before 09:20
                target_time = now.replace(hour=9, minute=20, second=0, microsecond=0)
                if now < target_time:
                    target_time = now.replace(hour=9, minute=15, second=0, microsecond=0)
                else:
                    minute = (now.minute // 5) * 5
                    target_time = now.replace(minute=minute, second=0, microsecond=0)
                from_time = target_time - timedelta(minutes=5)
                to_time = target_time
                range_from = int(from_time.timestamp())
                range_to = int(to_time.timestamp())
                data = {
                    "symbol": symbol,
                    "resolution": "5",
                    "date_format": "0",
                    "range_from": range_from,
                    "range_to": range_to,
                    "cont_flag": "1"
                }
                candles = self.fyers.history(data)
                if candles.get('s') == 'ok' and candles.get('candles'):
                    c = candles['candles'][-1]
                    o, h, l, cl = c[1], c[2], c[3], c[4]
                    candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                    self.log_info(f"[FALLBACK] Using available 5-min option candle for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                    return (o, h, l, cl, candle_time)
                else:
                    self.log_info(f"[RETRY] No 5-min option candle data returned for {symbol} from Fyers (attempt {attempt+1}/5). Retrying...")
                    time.sleep(2)
            except Exception as e:
                self.log_info(f"[RETRY] Error fetching 5-min option candle for {symbol} (attempt {attempt+1}/5): {e}")
                time.sleep(2)
        # As a last resort, try to fetch the most recent 5-min candle for today
        try:
            today = now.replace(hour=0, minute=0, second=0, microsecond=0)
            data = {
                "symbol": symbol,
                "resolution": "5",
                "date_format": "0",
                "range_from": int(today.timestamp()),
                "range_to": int(now.timestamp()),
                "cont_flag": "1"
            }
            candles = self.fyers.history(data)
            if candles.get('s') == 'ok' and candles.get('candles'):
                c = candles['candles'][-1]
                o, h, l, cl = c[1], c[2], c[3], c[4]
                candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                self.log_info(f"[FALLBACK-LAST] Using most recent 5-min option candle for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                return (o, h, l, cl, candle_time)
            else:
                self.log_info(f"[ERROR] No 5-min option candle data available for {symbol} even after all fallbacks.")
                return None
        except Exception as e:
            self.log_info(f"[ERROR] Final fallback error fetching 5-min option candle for {symbol}: {e}")
            return None

    def get_atm_option_symbol(self, spot, option_type, index_name):
        """Generate ATM option symbol with proper formatting"""
        try:
            # Check if BANKNIFTY options are enabled
            if 'BANK' in index_name.upper():
                banknifty_options_enabled = self.config.get('strategy', {}).get('banknifty_options_enabled', False)
                if not banknifty_options_enabled:
                    self.log_info(f"BANKNIFTY options not enabled in config - skipping {option_type} for {index_name}")
                    return None
            
            # Calculate step size and ATM strike
            step = 100 if 'BANK' in index_name.upper() else 50
            strike = round(spot / step) * step
            
            # Calculate expiry based on index type
            today = datetime.now(self.ist)
            
            if 'BANK' in index_name.upper():
                # BANKNIFTY: Always use next available expiry from Fyers option chain
                from src.banknifty_symbol_helper import get_next_banknifty_expiry, get_banknifty_option_symbol
                expiry_date = get_next_banknifty_expiry(today)
                underlying = 'BANKNIFTY'
                strike = round(spot / 100) * 100
                try:
                    symbol = get_banknifty_option_symbol(int(strike), option_type, expiry_date.date())
                    self.log_info(f"Selected BANKNIFTY {option_type} symbol from Fyers option chain: {symbol}")
                    return symbol
                except Exception as e:
                    self.log_info(f"[ERROR] BANKNIFTY option chain lookup failed: {e}")
                    # Fallback to formatter if option chain fails
                    from src.symbol_formatter import generate_option_symbol
                    symbol = generate_option_symbol(underlying, expiry_date.date(), int(strike), option_type)
                    self.log_info(f"Fallback BANKNIFTY symbol: {symbol}")
                    return symbol
            else:
                # NIFTY: Weekly options expire on Thursday  
                target_weekday = 3  # Thursday
                days_to_expiry = (target_weekday - today.weekday()) % 7
                if days_to_expiry == 0:  # Today is Thursday
                    if today.hour > 15 or (today.hour == 15 and today.minute >= 30):
                        days_to_expiry = 7  # Next Thursday after market close
                
                # For Oct 10, 2025, use Oct 14 expiry (which works based on logs)
                if today.date() <= datetime(2025, 10, 14).date():
                    expiry_date = datetime(2025, 10, 14, tzinfo=self.ist)
                else:
                    expiry_date = today + timedelta(days=days_to_expiry)
            
            # ...existing code for NIFTY only...
            # For NIFTY, use Fyers option chain to get exact symbol
            from src.nifty_symbol_helper import get_nifty_atm_option_symbol
            try:
                symbol = get_nifty_atm_option_symbol(spot, expiry_date.strftime('%d-%m-%Y'), option_type)
                if symbol:
                    self.log_info(f"Selected NIFTY {option_type} symbol from Fyers option chain: {symbol}")
                    return symbol
                else:
                    self.log_info(f"[ERROR] NIFTY option chain lookup failed, falling back to formatter.")
                    symbol = generate_option_symbol('NIFTY', expiry_date.date(), int(strike), option_type)
                    self.log_info(f"Fallback NIFTY symbol: {symbol}")
                    return symbol
            except Exception as e:
                self.log_info(f"[ERROR] NIFTY option chain lookup exception: {e}")
                symbol = generate_option_symbol('NIFTY', expiry_date.date(), int(strike), option_type)
                self.log_info(f"Fallback NIFTY symbol: {symbol}")
                return symbol
            
        except Exception as e:
            self.log_info(f"[ERROR] Failed to generate option symbol: {e}")
            return None
        # Convert to Fyers format
        return convert_option_symbol_format(symbol)

    def get_ltp(self, symbol):
        if self.simulation and not self.paper_trading:
            return 100
            
        # Use enhanced LTP method from DataFetcher if available
        if self.data_fetcher:
            try:
                ltp = self.data_fetcher.get_ltp_enhanced(symbol)
                if ltp is not None:
                    return ltp
            except Exception as e:
                self.log_info(f"[ERROR] DataFetcher LTP method failed: {e}")
        
        # Fall back to original method
        try:
            return get_ltp(self.fyers, symbol)
        except Exception as e:
            self.log_info(f"Error fetching LTP for {symbol}: {e}")
            return None

    def setup_websocket(self, symbols):
        # Temporarily disable WebSocket due to API parameter issues
        websocket_enabled = self.config.get('strategy', {}).get('enable_websocket', False)
        if not websocket_enabled:
            self.log_info("WebSocket disabled in configuration - using polling for live prices")
            return
            
        def ws_handler(symbol, key, value, tick_data):
            if key == 'ltp':
                self.live_prices[symbol] = float(value)
        try:
            self.data_socket = start_market_data_websocket(symbols=symbols, callback_handler=ws_handler)
            if self.data_socket:
                self.log_info(f"WebSocket subscription successful for: {symbols}")
            else:
                self.log_info("WebSocket subscription failed.")
        except Exception as e:
            self.log_info(f"WebSocket setup error: {e}")

    def monitor_breakout(self, symbol, ce_symbol, pe_symbol, ce_breakout, pe_breakout, qty, index_name, entry_buffer=2):
        self.log_info(f"Monitoring {symbol} for breakout. CE: {ce_symbol} ({ce_breakout}), PE: {pe_symbol} ({pe_breakout})")
        symbols_to_subscribe = [ce_symbol, pe_symbol]
        if not self.simulation or self.paper_trading:
            self.setup_websocket(symbols_to_subscribe)
        breakout_taken = False
        start_time = time.time()
        max_monitor_time = 60 * 60  # 1 hour max
        while not breakout_taken and (time.time() - start_time < max_monitor_time):
            for opt_symbol, breakout_level, opt_type in [
                (ce_symbol, ce_breakout, 'CE'),
                (pe_symbol, pe_breakout, 'PE')
            ]:
                # Do NOT fetch option OHLC at 9:20 here; just monitor LTP for breakout
                if self.simulation and not self.paper_trading:
                    ltp = breakout_level  # Simulate immediate breakout
                else:
                    ltp = self.live_prices.get(opt_symbol) or self.get_ltp(opt_symbol)
                # Check if LTP has broken above the breakout level
                if ltp is not None and ltp >= breakout_level:
                    # Check if entry price is not too far above breakout level (risk management)
                    max_premium_pct = self.config.get('strategy', {}).get('max_entry_premium_pct', 5)
                    premium_over_breakout = ((ltp - breakout_level) / breakout_level) * 100
                    
                    if premium_over_breakout > max_premium_pct:
                        self.log_info(f"WARNING: BREAKOUT DETECTED but ENTRY TOO RISKY!")
                        self.log_info(f"   {opt_type} LTP: {ltp} | Breakout: {breakout_level}")
                        self.log_info(f"   Premium over breakout: {premium_over_breakout:.1f}% (max allowed: {max_premium_pct}%)")
                        self.log_info(f"   Skipping entry to avoid overpriced trade")
                        # Continue monitoring for better entry or timeout
                        continue
                    
                    self.log_info(f"*** BREAKOUT DETECTED! {opt_type} option {opt_symbol} ***")
                    self.log_info(f"   Current LTP: {ltp} | Breakout Level: {breakout_level}")
                    self.log_info(f"   Premium over breakout: {premium_over_breakout:.1f}% (within {max_premium_pct}% limit)")
                    self.log_info(f"   Executing BUY order for {qty} lots...")
                    self.execute_trade(opt_symbol, ltp, qty, 'BUY', index_name)
                    breakout_taken = True
                    break
                else:
                    # Log current monitoring status every 30 seconds
                    if int(time.time()) % 30 == 0:
                        if ltp is not None:
                            self.log_info(f"Monitoring: {opt_type} {ltp:.2f} | Need: {breakout_level:.2f} | Gap: {(breakout_level - ltp):.2f}")
            time.sleep(2)  # Poll every 2s (30 calls/min)  Fyers limit: 200/min
        if not breakout_taken:
            self.log_info(f"No breakout detected for {symbol} within monitoring window.")

    def execute_trade(self, symbol, entry_price, lots, side, index_name):
        # Convert lots to quantity
        if 'NIFTY' in index_name and 'BANK' not in index_name:
            quantity = lots * 75  # NIFTY lot size
        else:
            quantity = lots * 35  # BANKNIFTY lot size
            
        sl = entry_price - self.sl_points
        target = entry_price + self.target_points
        entry_time = datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S')
        
        if self.paper_trading:
            self.log_info(f"[PAPER TRADE] {side} {symbol} - {lots} lots ({quantity} qty) at {entry_price}")
            self.log_info(f"   Stop Loss: {sl} | Target: {target}")
        else:
            self.log_info(f"Trade executed: {side} {symbol} - {lots} lots ({quantity} qty) at {entry_price}")
            self.log_info(f"   Stop Loss: {sl} | Target: {target}")
        
        self.log_trade(symbol, entry_price, quantity, side, 'BREAKOUT', entry_time)
        self.manage_position(symbol, entry_price, quantity, sl, target, side, entry_time, index_name)

    def manage_position(self, symbol, entry, qty, sl, target, side, entry_time, index_name):
        max_holding_minutes = 60
        trailing_sl = sl
        exit_reason = None
        max_up = float('-inf')  # Maximum unrealized profit
        max_down = float('inf') # Maximum drawdown (largest unrealized loss)
        for minute in range(max_holding_minutes * 60):  # every second
            if self.simulation and not self.paper_trading:
                ltp = entry + self.target_points  # Simulate target hit
            else:
                ltp = self.get_ltp(symbol)
            pnl = (ltp - entry) * qty if side == 'BUY' else (entry - ltp) * qty
            pnl_pct = ((ltp - entry) / entry) * 100 if entry else 0
            # Track max_up and max_down
            if pnl > max_up:
                max_up = pnl
            if pnl < max_down:
                max_down = pnl
            self.log_info(f"[MONITOR] {symbol} | Entry: {entry} | LTP: {ltp} | PnL: {pnl:.2f} | SL: {sl} | Trailing SL: {trailing_sl} | PnL%: {pnl_pct:.2f} | MaxUp: {max_up:.2f} | MaxDown: {max_down:.2f}")
            if ltp <= trailing_sl:
                exit_reason = 'STOPLOSS'
                exit_price = trailing_sl
                break
            elif ltp >= target:
                exit_reason = 'TARGET'
                exit_price = target
                break
            # Trailing SL logic
            if ltp > entry and ltp - entry > self.sl_points:
                new_trailing = ltp - self.sl_points
                if new_trailing > trailing_sl:
                    self.log_info(f"Trailing SL moved up to {new_trailing}")
                    trailing_sl = new_trailing
            time.sleep(2)  # Poll every 2s (30 calls/min)  Fyers limit: 200/min
        else:
            exit_reason = 'TIME_EXIT'
            exit_price = ltp
        exit_time = datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S')
        self.log_info(f"Exiting {symbol} at {exit_price} due to {exit_reason} | MaxUp: {max_up:.2f} | MaxDown: {max_down:.2f}")
        self.log_trade(symbol, exit_price, qty, 'SELL', exit_reason, exit_time)

    def log_trade(self, symbol, price, qty, side, reason, time_str):
        row = f'{time_str},{symbol},{side},{price},{qty},{reason}\n'
        with open(self.log_file, 'a') as f:
            f.write(row)
        self.logger.info(f"Trade logged: {row.strip()}")

    def _append_final_row_with_format(self, excel_file, csv_file, final_row, columns):
        """Append a final_row to excel_file with proper formatting (bold headers, aligned columns, frozen pane).
        Also append to CSV. Creates fresh file if needed.
        """
        import pandas as pd
        import csv
        import os
        
        # Round all values to 2 decimals if float
        def round2(val):
            try:
                return round(float(val), 2)
            except Exception:
                return val
        
        final_row_rounded = [round2(x) for x in final_row]
        
        # Write to Excel with proper formatting
        try:
            import openpyxl
            from openpyxl import load_workbook, Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Load or create workbook
            if os.path.exists(excel_file):
                wb = load_workbook(excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Trade History"
                
                # Add headers with formatting
                ws.append(columns)
                
                # Style headers: Bold, centered, with background color
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Freeze first row
                ws.freeze_panes = "A2"
            
            # Append data row
            ws.append(final_row_rounded)
            
            # Format data cells with center alignment for better readability
            data_row = ws.max_row
            data_alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.alignment = data_alignment
            
            # Auto-adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ''
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 3, 50)  # Max width 50 to avoid huge columns
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for better visibility
            ws.row_dimensions[1].height = 20  # Header row
            ws.row_dimensions[data_row].height = 18  # Data row
            
            # Save workbook
            wb.save(excel_file)
            self.log_info(f"Trade data saved to Excel: {excel_file}")
            
        except Exception as e:
            self.log_info(f"[ERROR] Failed to write Excel file {excel_file}: {e}")
            import traceback
            self.log_info(f"Traceback: {traceback.format_exc()}")
        
        # Write to CSV (append mode, add headers only if new)
        try:
            file_exists = os.path.exists(csv_file)
            with open(csv_file, 'a', newline='', encoding='utf-8') as cf:
                writer = csv.writer(cf)
                if not file_exists or os.stat(csv_file).st_size == 0:
                    writer.writerow(columns)
                writer.writerow([str(x) if x is not None else '' for x in final_row_rounded])
            self.log_info(f"Trade data saved to CSV: {csv_file}")
        except Exception as e:
            self.log_info(f"[ERROR] Failed to write CSV file {csv_file}: {e}")

    def place_bracket_order(self, symbol, entry_price, qty, breakout_level, index_name):
        """
        Simulate placing a bracket order. In live mode, integrate with broker API here.
        Returns a simulated order ID string.
        """
        if self.simulation or self.paper_trading:
            # Simulate order placement
            order_id = f"SIM-{symbol}-{int(entry_price)}-{qty}"
            self.log_info(f"[SIMULATION] Placed bracket order for {symbol} @ {entry_price} qty={qty} (breakout={breakout_level})")
            return order_id
        else:
            # TODO: Integrate with broker API for live trading
            self.log_info(f"[LIVE] Placing real bracket order for {symbol} @ {entry_price} qty={qty} (breakout={breakout_level})")
            # Example: return broker.place_bracket_order(...)
            return None

    def get_order_status(self, order_id):
        """
        Simulate order status for paper/simulation mode. Always returns 'PENDING' for now.
        Extend this logic to simulate fills/cancellations as needed.
        """
        # In a real implementation, this would check broker API or paper_orders dict
        if hasattr(self, 'paper_orders') and order_id in self.paper_orders:
            return self.paper_orders[order_id].get('status', 'PENDING')
        return 'PENDING'

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--simulate', action='store_true', help='Run in simulation mode (dummy data)')
    parser.add_argument('--paper', action='store_true', help='Run in paper trading mode (real data, no real trades)')
    args = parser.parse_args()
    strategy = Breakout5MinStrategy(simulation=args.simulate, paper_trading=args.paper)
    strategy.run()
    # Archive the log file at the end of the session
    archive_strategy_log()

    def wait_for_market_open(self):
        now = datetime.now(self.ist)
        market_open = now.replace(hour=9, minute=15, second=0, microsecond=0)
        if now >= market_open:
            # After 09:15, but always want to use 09:15-09:20 window for breakout
            self.skip_time_rule = False  # Always use 09:15-09:20
            self.log_info("[USER] Script started after 09:15. Will attempt to fetch 09:15-09:20 OHLC for breakout.")
            return
        while True:
            now = datetime.now(self.ist).time()
            if now >= datetime.strptime('09:15', '%H:%M').time():
                break
            self.log_info('Waiting for market to open (09:15 IST)...')
            time.sleep(30)

    def wait_until_920(self):
        now = datetime.now(self.ist)
        target = now.replace(hour=9, minute=20, second=0, microsecond=0)
        if now < target:
            seconds = (target - now).total_seconds()
            self.log_info(f'Waiting {int(seconds)} seconds until 9:20 IST for first 5-min candle to form...')
            time.sleep(seconds)
            self.log_info('Reached 9:20 IST. Proceeding to fetch first 5-min candle.')

    def collect_live_5min_ohlc(self, symbol):
        # Deprecated: No longer used. Always use Fyers historical API for index OHLC.
        self.log_info(f"[SKIP] collect_live_5min_ohlc is disabled. Using Fyers historical API only for {symbol}.")
        return None

    def fetch_5min_candle(self, symbol):
        # Always use Fyers historical API to fetch 5-min OHLC for index at 09:20 IST
        if self.simulation and not self.paper_trading:
            return (20000, 20020, 19980, 20010, '09:15')
            
        # Use the new DataFetcher for more reliable candle data
        candle_data = self.data_fetcher.get_first_5min_candle(symbol)
        if candle_data:
            o, h, l, cl, candle_time = candle_data
            self.log_info(f"[ENHANCED] 5-min OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
            return candle_data
            
        # If DataFetcher fails, fall back to the old method
        self.log_info(f"[FALLBACK] DataFetcher failed for {symbol}, trying old method")
        
        from datetime import datetime, timedelta
        ist = pytz.timezone('Asia/Kolkata')
        now = datetime.now(ist)
        # Target 09:15-09:20 candle for breakout
        target_time = now.replace(hour=9, minute=20, second=0, microsecond=0)
        if now < target_time:
            # If before 9:20, wait until 9:20
            wait_sec = (target_time - now).total_seconds()
            if wait_sec > 0:
                self.log_info(f"Waiting {int(wait_sec)} seconds until 9:20 IST for first 5-min candle to form...")
                time.sleep(wait_sec)
            now = datetime.now(ist)
        # Retry logic for Fyers API
        for attempt in range(5):
            try:
                from_time = target_time - timedelta(minutes=5)
                to_time = target_time
                range_from = int(from_time.timestamp())
                range_to = int(to_time.timestamp())
                data = {
                    "symbol": symbol,
                    "resolution": "5",
                    "date_format": "0",  # Use 0 for epoch timestamps
                    "range_from": range_from,
                    "range_to": range_to,
                    "cont_flag": "1"
                }
                candles = self.fyers.history(data)
                if candles.get('s') == 'ok' and candles.get('candles'):
                    c = candles['candles'][-1]
                    o, h, l, cl = c[1], c[2], c[3], c[4]
                    candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                    self.log_info(f"[API] 5-min OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                    return (o, h, l, cl, candle_time)
                else:
                    self.log_info(f"[RETRY] No 5-min candle data returned for {symbol} from Fyers (attempt {attempt+1}/5). Retrying...")
                    time.sleep(2)
            except Exception as e:
                self.log_info(f"[RETRY] Error fetching 5-min candle for {symbol} (attempt {attempt+1}/5): {e}")
                time.sleep(2)
        # Fallback: try to fetch the most recent 5-min candle for today
        try:
            today = now.replace(hour=0, minute=0, second=0, microsecond=0)
            data = {
                "symbol": symbol,
                "resolution": "5",
                "date_format": "0",
                "range_from": int(today.timestamp()),
                "range_to": int(now.timestamp()),
                "cont_flag": "1"
            }
            candles = self.fyers.history(data)
            if candles.get('s') == 'ok' and candles.get('candles'):
                c = candles['candles'][-1]
                o, h, l, cl = c[1], c[2], c[3], c[4]
                candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                self.log_info(f"[FALLBACK-LAST] Using most recent 5-min candle for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                return (o, h, l, cl, candle_time)
            else:
                self.log_info(f"[ERROR] No 5-min candle data available for {symbol} even after all fallbacks.")
                return None
        except Exception as e:
            self.log_info(f"[ERROR] Final fallback error fetching 5-min candle for {symbol}: {e}")
            return None

    def fetch_option_ohlc(self, symbol):
        # Fetch 5-min OHLC for option symbol using Fyers historical API with retry/fallback
        if self.simulation and not self.paper_trading:
            return (100, 106, 99, 105, '09:20')
            
        # Use the new DataFetcher for more reliable option data
        if self.data_fetcher:
            try:
                candle_data = self.data_fetcher.get_first_5min_candle(symbol)
                if candle_data:
                    o, h, l, cl, candle_time = candle_data
                    self.log_info(f"[ENHANCED] 5-min option OHLC for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                    return candle_data
            except Exception as e:
                self.log_info(f"[ERROR] DataFetcher failed for option {symbol}: {e}")
        
        # If DataFetcher fails, fall back to the old method
        ist = pytz.timezone('Asia/Kolkata')
        now = datetime.now(ist)
        # Try up to 5 times with 2s delay to allow for Fyers data lag
        for attempt in range(5):
            try:
                # Try to get the 09:15-09:20 candle, else fallback to latest available before 09:20
                target_time = now.replace(hour=9, minute=20, second=0, microsecond=0)
                if now < target_time:
                    target_time = now.replace(hour=9, minute=15, second=0, microsecond=0)
                else:
                    minute = (now.minute // 5) * 5
                    target_time = now.replace(minute=minute, second=0, microsecond=0)
                from_time = target_time - timedelta(minutes=5)
                to_time = target_time
                range_from = int(from_time.timestamp())
                range_to = int(to_time.timestamp())
                data = {
                    "symbol": symbol,
                    "resolution": "5",
                    "date_format": "0",
                    "range_from": range_from,
                    "range_to": range_to,
                    "cont_flag": "1"
                }
                candles = self.fyers.history(data)
                if candles.get('s') == 'ok' and candles.get('candles'):
                    c = candles['candles'][-1]
                    o, h, l, cl = c[1], c[2], c[3], c[4]
                    candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                    self.log_info(f"[FALLBACK] Using available 5-min option candle for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                    return (o, h, l, cl, candle_time)
                else:
                    self.log_info(f"[RETRY] No 5-min option candle data returned for {symbol} from Fyers (attempt {attempt+1}/5). Retrying...")
                    time.sleep(2)
            except Exception as e:
                self.log_info(f"[RETRY] Error fetching 5-min option candle for {symbol} (attempt {attempt+1}/5): {e}")
                time.sleep(2)
        # As a last resort, try to fetch the most recent 5-min candle for today
        try:
            today = now.replace(hour=0, minute=0, second=0, microsecond=0)
            data = {
                "symbol": symbol,
                "resolution": "5",
                "date_format": "0",
                "range_from": int(today.timestamp()),
                "range_to": int(now.timestamp()),
                "cont_flag": "1"
            }
            candles = self.fyers.history(data)
            if candles.get('s') == 'ok' and candles.get('candles'):
                c = candles['candles'][-1]
                o, h, l, cl = c[1], c[2], c[3], c[4]
                candle_time = datetime.fromtimestamp(c[0], ist).strftime('%H:%M')
                self.log_info(f"[FALLBACK-LAST] Using most recent 5-min option candle for {symbol}: O={o}, H={h}, L={l}, C={cl}, Time={candle_time}")
                return (o, h, l, cl, candle_time)
            else:
                self.log_info(f"[ERROR] No 5-min option candle data available for {symbol} even after all fallbacks.")
                return None
        except Exception as e:
            self.log_info(f"[ERROR] Final fallback error fetching 5-min option candle for {symbol}: {e}")
            return None

    def get_atm_option_symbol(self, spot, option_type, index_name):
        """Generate ATM option symbol with proper formatting"""
        try:
            # Check if BANKNIFTY options are enabled
            if 'BANK' in index_name.upper():
                banknifty_options_enabled = self.config.get('strategy', {}).get('banknifty_options_enabled', False)
                if not banknifty_options_enabled:
                    self.log_info(f"BANKNIFTY options not enabled in config - skipping {option_type} for {index_name}")
                    return None
            
            # Calculate step size and ATM strike
            step = 100 if 'BANK' in index_name.upper() else 50
            strike = round(spot / step) * step
            
            # Calculate expiry based on index type
            today = datetime.now(self.ist)
            
            if 'BANK' in index_name.upper():
                # BANKNIFTY: Always use next available expiry from Fyers option chain
                from src.banknifty_symbol_helper import get_next_banknifty_expiry, get_banknifty_option_symbol
                expiry_date = get_next_banknifty_expiry(today)
                underlying = 'BANKNIFTY'
                strike = round(spot / 100) * 100
                try:
                    symbol = get_banknifty_option_symbol(int(strike), option_type, expiry_date.date())
                    self.log_info(f"Selected BANKNIFTY {option_type} symbol from Fyers option chain: {symbol}")
                    return symbol
                except Exception as e:
                    self.log_info(f"[ERROR] BANKNIFTY option chain lookup failed: {e}")
                    # Fallback to formatter if option chain fails
                    from src.symbol_formatter import generate_option_symbol
                    symbol = generate_option_symbol(underlying, expiry_date.date(), int(strike), option_type)
                    self.log_info(f"Fallback BANKNIFTY symbol: {symbol}")
                    return symbol
            else:
                # NIFTY: Weekly options expire on Thursday  
                target_weekday = 3  # Thursday
                days_to_expiry = (target_weekday - today.weekday()) % 7
                if days_to_expiry == 0:  # Today is Thursday
                    if today.hour > 15 or (today.hour == 15 and today.minute >= 30):
                        days_to_expiry = 7  # Next Thursday after market close
                
                # For Oct 10, 2025, use Oct 14 expiry (which works based on logs)
                if today.date() <= datetime(2025, 10, 14).date():
                    expiry_date = datetime(2025, 10, 14, tzinfo=self.ist)
                else:
                    expiry_date = today + timedelta(days=days_to_expiry)
            
            # ...existing code for NIFTY only...
            # For NIFTY, use Fyers option chain to get exact symbol
            from src.nifty_symbol_helper import get_nifty_atm_option_symbol
            try:
                symbol = get_nifty_atm_option_symbol(spot, expiry_date.strftime('%d-%m-%Y'), option_type)
                if symbol:
                    self.log_info(f"Selected NIFTY {option_type} symbol from Fyers option chain: {symbol}")
                    return symbol
                else:
                    self.log_info(f"[ERROR] NIFTY option chain lookup failed, falling back to formatter.")
                    symbol = generate_option_symbol('NIFTY', expiry_date.date(), int(strike), option_type)
                    self.log_info(f"Fallback NIFTY symbol: {symbol}")
                    return symbol
            except Exception as e:
                self.log_info(f"[ERROR] NIFTY option chain lookup exception: {e}")
                symbol = generate_option_symbol('NIFTY', expiry_date.date(), int(strike), option_type)
                self.log_info(f"Fallback NIFTY symbol: {symbol}")
                return symbol
            
        except Exception as e:
            self.log_info(f"[ERROR] Failed to generate option symbol: {e}")
            return None
        # Convert to Fyers format
        return convert_option_symbol_format(symbol)

    def get_ltp(self, symbol):
        if self.simulation and not self.paper_trading:
            return 100
            
        # Use enhanced LTP method from DataFetcher if available
        if self.data_fetcher:
            try:
                ltp = self.data_fetcher.get_ltp_enhanced(symbol)
                if ltp is not None:
                    return ltp
            except Exception as e:
                self.log_info(f"[ERROR] DataFetcher LTP method failed: {e}")
        
        # Fall back to original method
        try:
            return get_ltp(self.fyers, symbol)
        except Exception as e:
            self.log_info(f"Error fetching LTP for {symbol}: {e}")
            return None

    def setup_websocket(self, symbols):
        # Temporarily disable WebSocket due to API parameter issues
        websocket_enabled = self.config.get('strategy', {}).get('enable_websocket', False)
        if not websocket_enabled:
            self.log_info("WebSocket disabled in configuration - using polling for live prices")
            return
            
        def ws_handler(symbol, key, value, tick_data):
            if key == 'ltp':
                self.live_prices[symbol] = float(value)
        try:
            self.data_socket = start_market_data_websocket(symbols=symbols, callback_handler=ws_handler)
            if self.data_socket:
                self.log_info(f"WebSocket subscription successful for: {symbols}")
            else:
                self.log_info("WebSocket subscription failed.")
        except Exception as e:
            self.log_info(f"WebSocket setup error: {e}")

    def monitor_breakout(self, symbol, ce_symbol, pe_symbol, ce_breakout, pe_breakout, qty, index_name, entry_buffer=2):
        self.log_info(f"Monitoring {symbol} for breakout. CE: {ce_symbol} ({ce_breakout}), PE: {pe_symbol} ({pe_breakout})")
        symbols_to_subscribe = [ce_symbol, pe_symbol]
        if not self.simulation or self.paper_trading:
            self.setup_websocket(symbols_to_subscribe)
        breakout_taken = False
        start_time = time.time()
        max_monitor_time = 60 * 60  # 1 hour max
        while not breakout_taken and (time.time() - start_time < max_monitor_time):
            for opt_symbol, breakout_level, opt_type in [
                (ce_symbol, ce_breakout, 'CE'),
                (pe_symbol, pe_breakout, 'PE')
            ]:
                # Do NOT fetch option OHLC at 9:20 here; just monitor LTP for breakout
                if self.simulation and not self.paper_trading:
                    ltp = breakout_level  # Simulate immediate breakout
                else:
                    ltp = self.live_prices.get(opt_symbol) or self.get_ltp(opt_symbol)
                # Check if LTP has broken above the breakout level
                if ltp is not None and ltp >= breakout_level:
                    # Check if entry price is not too far above breakout level (risk management)
                    max_premium_pct = self.config.get('strategy', {}).get('max_entry_premium_pct', 5)
                    premium_over_breakout = ((ltp - breakout_level) / breakout_level) * 100
                    
                    if premium_over_breakout > max_premium_pct:
                        self.log_info(f"WARNING: BREAKOUT DETECTED but ENTRY TOO RISKY!")
                        self.log_info(f"   {opt_type} LTP: {ltp} | Breakout: {breakout_level}")
                        self.log_info(f"   Premium over breakout: {premium_over_breakout:.1f}% (max allowed: {max_premium_pct}%)")
                        self.log_info(f"   Skipping entry to avoid overpriced trade")
                        # Continue monitoring for better entry or timeout
                        continue
                    
                    self.log_info(f"*** BREAKOUT DETECTED! {opt_type} option {opt_symbol} ***")
                    self.log_info(f"   Current LTP: {ltp} | Breakout Level: {breakout_level}")
                    self.log_info(f"   Premium over breakout: {premium_over_breakout:.1f}% (within {max_premium_pct}% limit)")
                    self.log_info(f"   Executing BUY order for {qty} lots...")
                    self.execute_trade(opt_symbol, ltp, qty, 'BUY', index_name)
                    breakout_taken = True
                    break
                else:
                    # Log current monitoring status every 30 seconds
                    if int(time.time()) % 30 == 0:
                        if ltp is not None:
                            self.log_info(f"Monitoring: {opt_type} {ltp:.2f} | Need: {breakout_level:.2f} | Gap: {(breakout_level - ltp):.2f}")
            time.sleep(2)  # Poll every 2s (30 calls/min)  Fyers limit: 200/min
        if not breakout_taken:
            self.log_info(f"No breakout detected for {symbol} within monitoring window.")

    def execute_trade(self, symbol, entry_price, lots, side, index_name):
        # Convert lots to quantity
        if 'NIFTY' in index_name and 'BANK' not in index_name:
            quantity = lots * 75  # NIFTY lot size
        else:
            quantity = lots * 35  # BANKNIFTY lot size
            
        sl = entry_price - self.sl_points
        target = entry_price + self.target_points
        entry_time = datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S')
        
        if self.paper_trading:
            self.log_info(f"[PAPER TRADE] {side} {symbol} - {lots} lots ({quantity} qty) at {entry_price}")
            self.log_info(f"   Stop Loss: {sl} | Target: {target}")
        else:
            self.log_info(f"Trade executed: {side} {symbol} - {lots} lots ({quantity} qty) at {entry_price}")
            self.log_info(f"   Stop Loss: {sl} | Target: {target}")
        
        self.log_trade(symbol, entry_price, quantity, side, 'BREAKOUT', entry_time)
        self.manage_position(symbol, entry_price, quantity, sl, target, side, entry_time, index_name)

    def manage_position(self, symbol, entry, qty, sl, target, side, entry_time, index_name):
        max_holding_minutes = 60
        trailing_sl = sl
        exit_reason = None
        max_up = float('-inf')  # Maximum unrealized profit
        max_down = float('inf') # Maximum drawdown (largest unrealized loss)
        for minute in range(max_holding_minutes * 60):  # every second
            if self.simulation and not self.paper_trading:
                ltp = entry + self.target_points  # Simulate target hit
            else:
                ltp = self.get_ltp(symbol)
            pnl = (ltp - entry) * qty if side == 'BUY' else (entry - ltp) * qty
            pnl_pct = ((ltp - entry) / entry) * 100 if entry else 0
            # Track max_up and max_down
            if pnl > max_up:
                max_up = pnl
            if pnl < max_down:
                max_down = pnl
            self.log_info(f"[MONITOR] {symbol} | Entry: {entry} | LTP: {ltp} | PnL: {pnl:.2f} | SL: {sl} | Trailing SL: {trailing_sl} | PnL%: {pnl_pct:.2f} | MaxUp: {max_up:.2f} | MaxDown: {max_down:.2f}")
            if ltp <= trailing_sl:
                exit_reason = 'STOPLOSS'
                exit_price = trailing_sl
                break
            elif ltp >= target:
                exit_reason = 'TARGET'
                exit_price = target
                break
            # Trailing SL logic
            if ltp > entry and ltp - entry > self.sl_points:
                new_trailing = ltp - self.sl_points
                if new_trailing > trailing_sl:
                    self.log_info(f"Trailing SL moved up to {new_trailing}")
                    trailing_sl = new_trailing
            time.sleep(2)  # Poll every 2s (30 calls/min)  Fyers limit: 200/min
        else:
            exit_reason = 'TIME_EXIT'
            exit_price = ltp
        exit_time = datetime.now(self.ist).strftime('%Y-%m-%d %H:%M:%S')
        self.log_info(f"Exiting {symbol} at {exit_price} due to {exit_reason} | MaxUp: {max_up:.2f} | MaxDown: {max_down:.2f}")
        self.log_trade(symbol, exit_price, qty, 'SELL', exit_reason, exit_time)

    def log_trade(self, symbol, price, qty, side, reason, time_str):
        row = f'{time_str},{symbol},{side},{price},{qty},{reason}\n'
        with open(self.log_file, 'a') as f:
            f.write(row)
        self.logger.info(f"Trade logged: {row.strip()}")

    def _append_final_row_with_format(self, excel_file, csv_file, final_row, columns):
        """Append a final_row to excel_file with proper formatting (bold headers, aligned columns, frozen pane).
        Also append to CSV. Creates fresh file if needed.
        """
        import pandas as pd
        import csv
        import os
        
        # Round all values to 2 decimals if float
        def round2(val):
            try:
                return round(float(val), 2)
            except Exception:
                return val
        
        final_row_rounded = [round2(x) for x in final_row]
        
        # Write to Excel with proper formatting
        try:
            import openpyxl
            from openpyxl import load_workbook, Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Load or create workbook
            if os.path.exists(excel_file):
                wb = load_workbook(excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Trade History"
                
                # Add headers with formatting
                ws.append(columns)
                
                # Style headers: Bold, centered, with background color
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Freeze first row
                ws.freeze_panes = "A2"
            
            # Append data row
            ws.append(final_row_rounded)
            
            # Format data cells with center alignment for better readability
            data_row = ws.max_row
            data_alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.alignment = data_alignment
            
            # Auto-adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ''
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = min(max_length + 3, 50)  # Max width 50 to avoid huge columns
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for better visibility
            ws.row_dimensions[1].height = 20  # Header row
            ws.row_dimensions[data_row].height = 18  # Data row
            
            # Save workbook
            wb.save(excel_file)
            self.log_info(f"Trade data saved to Excel: {excel_file}")
            
        except Exception as e:
            self.log_info(f"[ERROR] Failed to write Excel file {excel_file}: {e}")
            import traceback
            self.log_info(f"Traceback: {traceback.format_exc()}")
        
        # Write to CSV (append mode, add headers only if new)
        try:
            file_exists = os.path.exists(csv_file)
            with open(csv_file, 'a', newline='', encoding='utf-8') as cf:
                writer = csv.writer(cf)
                if not file_exists or os.stat(csv_file).st_size == 0:
                    writer.writerow(columns)
                writer.writerow([str(x) if x is not None else '' for x in final_row_rounded])
            self.log_info(f"Trade data saved to CSV: {csv_file}")
        except Exception as e:
            self.log_info(f"[ERROR] Failed to write CSV file {csv_file}: {e}")

    def place_bracket_order(self, symbol, entry_price, qty, breakout_level, index_name):
        """
        Simulate placing a bracket order. In live mode, integrate with broker API here.
        Returns a simulated order ID string.
        """
        if self.simulation or self.paper_trading:
            # Simulate order placement
            order_id = f"SIM-{symbol}-{int(entry_price)}-{qty}"
            self.log_info(f"[SIMULATION] Placed bracket order for {symbol} @ {entry_price} qty={qty} (breakout={breakout_level})")
            return order_id
        else:
            # TODO: Integrate with broker API for live trading
            self.log_info(f"[LIVE] Placing real bracket order for {symbol} @ {entry_price} qty={qty} (breakout={breakout_level})")
            # Example: return broker.place_bracket_order(...)
            return None

    def get_order_status(self, order_id):
        """
        Simulate order status for paper/simulation mode. Always returns 'PENDING' for now.
        Extend this logic to simulate fills/cancellations as needed.
        """
        # In a real implementation, this would check broker API or paper_orders dict
        if hasattr(self, 'paper_orders') and order_id in self.paper_orders:
            return self.paper_orders[order_id].get('status', 'PENDING')
        return 'PENDING'

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--simulate', action='store_true', help='Run in simulation mode (dummy data)')
    parser.add_argument('--paper', action='store_true', help='Run in paper trading mode (real data, no real trades)')
    args = parser.parse_args()
    strategy = Breakout5MinStrategy(simulation=args.simulate, paper_trading=args.paper)
    strategy.run()
